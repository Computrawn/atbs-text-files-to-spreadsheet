#!/usr/bin/env python3
# text_files_to_spreadsheet.py — An exercise in manipulating Excel files.
# For more information, see README.md

from pathlib import Path
import logging
import openpyxl
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.DEBUG,
    filename="logging.txt",
    format="%(asctime)s -  %(levelname)s -  %(message)s",
)

logging.disable(logging.CRITICAL)  # Note out to enable logging.


def find_files():
    """Find all txt files in designated folder
    and make a list of their file paths."""
    user_path = Path(input("Please type file path here: "))
    document_list = list(user_path.glob("*.txt"))
    logging.debug(document_list)
    return document_list


def pull_contents(file_list):
    """Convert contents of each line of each text file into a list of lists."""
    document_contents = []
    for item in file_list:
        with open(item, "rb") as f_contents:
            lines = f_contents.readlines()
            document_contents.append(lines)
    return document_contents


def write_excel(contents):
    """Create an Excel file with every txt files as its own column
    and each line of the txt file as a row."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    column_number = len(contents)

    for idx in range(column_number):
        column_letter = get_column_letter(idx + 1)
        logging.debug(column_letter)
        row_number = len(contents[idx])

        for item in range(row_number):
            sheet_info = f"{column_letter}{item + 1}"
            logging.debug(sheet_info)
            sheet[f"{sheet_info}"] = contents[idx][item]

    wb.save("text_to_spread.xlsx")


def main_func():
    """Executes prior functions in sequence."""
    doc_list = find_files()
    doc_contents = pull_contents(doc_list)
    logging.debug(doc_contents)
    write_excel(doc_contents)


main_func()
